import json
from pathlib import Path
from openpyxl import load_workbook

repo_root = Path(__file__).resolve().parent.parent
excel_path = repo_root / 'scripts' / 'catalog-data' / 'source' / 'Περιγραφές GUSTO.xlsx'
json_path = repo_root / 'public' / 'data' / 'products-grouped.json'
backup_path = json_path.with_suffix('.json.bak')

if not excel_path.exists():
    print('Excel not found:', excel_path)
    raise SystemExit(1)
if not json_path.exists():
    print('JSON not found:', json_path)
    raise SystemExit(1)

wb = load_workbook(excel_path, read_only=True, data_only=True)
sheet_name = None
for n in wb.sheetnames:
    if 'final' in n.lower():
        sheet_name = n
        break
if sheet_name is None:
    sheet_name = wb.sheetnames[0]
sh = wb[sheet_name]

# read rows as lists
rows = list(sh.values)

header_row_index = -1
for i, r in enumerate(rows):
    if not r:
        continue
    for c in r:
        if isinstance(c, str) and 'κωδ' in c.lower():
            header_row_index = i
            break
    if header_row_index != -1:
        break

if header_row_index == -1:
    print('Header not found in excel')
    raise SystemExit(1)

header = [str(x).strip() if x is not None else '' for x in rows[header_row_index]]

def find_header(names):
    for idx, h in enumerate(header):
        if not h:
            continue
        low = h.lower()
        for n in names:
            if n in low:
                return idx
    return -1

code_col = find_header(['κωδ','code'])
analy_col = find_header(['αναλυτική περιγραφή gr','αναλυτική περιγραφή','αναλυτική'])
analy_en_col = find_header(['αναλυτική περιγραφή en','analytical description en','english'])
tech_col = find_header(['τεχνικά χαρακτηριστικά','technical details','τεχνικά'])
tech_en_col = find_header(['technical details'])
care_col = find_header(['οδηγίες χρήσης','οδηγίες χρήσης & φροντίδας','use & care','οδηγίες'])
care_en_col = find_header(['use & care','use & care instructions','care'])

mapping = {}
for r in rows[header_row_index+1:]:
    if not r:
        continue
    code = ''
    if code_col != -1 and code_col < len(r) and r[code_col] is not None:
        code = str(r[code_col]).strip()
    if not code:
        continue
    def cell_text(idx):
        if idx == -1 or idx >= len(r) or r[idx] is None:
            return ''
        return str(r[idx]).strip()
    analy = cell_text(analy_col)
    analy_en = cell_text(analy_en_col)
    tech = cell_text(tech_col)
    tech_en = cell_text(tech_en_col)
    care = cell_text(care_col)
    care_en = cell_text(care_en_col)
    mapping[code] = {'analy':analy,'analy_en':analy_en,'tech':tech,'tech_en':tech_en,'care':care,'care_en':care_en}

data = json.loads(json_path.read_text(encoding='utf8'))
backup_path.write_text(json_path.read_text(encoding='utf8'), encoding='utf8')

updated = 0
for product in data.get('products', []):
    for size in product.get('sizes', []):
        for variant in size.get('variants', []):
            code = str(variant.get('code','')).strip()
            if code in mapping:
                e = mapping[code]
                changed = False
                if e['analy'] and variant.get('excel_ar') != e['analy']:
                    variant['excel_ar'] = e['analy']; changed = True
                if e['analy_en'] and variant.get('excel_en') != e['analy_en']:
                    variant['excel_en'] = e['analy_en']; changed = True
                if e['tech'] and variant.get('excel_tech_gr') != e['tech']:
                    variant['excel_tech_gr'] = e['tech']; changed = True
                if e['tech_en'] and variant.get('excel_tech_en') != e['tech_en']:
                    variant['excel_tech_en'] = e['tech_en']; changed = True
                if e['care'] and variant.get('excel_care_gr') != e['care']:
                    variant['excel_care_gr'] = e['care']; changed = True
                if e['care_en'] and variant.get('excel_care_en') != e['care_en']:
                    variant['excel_care_en'] = e['care_en']; changed = True
                if changed:
                    updated += 1

json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf8')
print('Done. Updated entries:', updated)
print('Backup saved to', backup_path)
