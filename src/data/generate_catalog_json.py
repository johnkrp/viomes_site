#!/usr/bin/env python3
"""
Generate:
  - products-grouped.json
  - additional-images.json

from the VIOMES Excel catalog workbook.

Usage:
  python src/data/generate_catalog_json.py
  python src/data/generate_catalog_json.py --xlsx "src/data/ΚΑΤΑΣΤΑΣΗ ΕΙΔΩΝ ΚΑΙ ΠΕΡΙΓΡΑΦΕΣ SITE.xlsx"
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import openpyxl


# Zero-based column indices in the worksheet.
COL_GROUP_MARKER = 22      # W
COL_FAMILY_MARKER = 23     # X
COL_CODE = 3               # D
COL_DESCRIPTION = 5        # F
COL_COLOR = 8              # I
COL_PACK = 11              # L
COL_CATEGORY = 29          # AD
COL_TITLE_GR = 40          # AO (ΕΛΛΗΝΙΚΗ ΠΕΡΙΓΡΑΦΗ SITE)
COL_TITLE_EN_SLUG = 41     # AP (ΑΓΓΛΙΚΗ ΠΕΡΙΓΡΑΦΗ ΓΙΑ ΣΚΡΟΥΤΖ)
COL_EXCEL_AR = 43          # AR (ΚΕΙΜΕΝΟ SITE ΕΛΛΗΝΙΚΑ)
COL_PACKSHOT = 51          # AZ
COL_LIFESTYLE_START = 52   # BA..BD
COL_LIFESTYLE_END = 55
GREEN_FILL_HEX_SUFFIX = "C6EFCE"


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.lower() == "nan":
        return ""
    return text


def normalize_ascii_slug(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text.lower())
    without_marks = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    ascii_only = without_marks.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_only).strip("-")
    return slug


def extract_group_code(group_root: str, fallback_code: str) -> str:
    match = re.search(r"\d+", group_root or "")
    if match:
        return match.group(0)
    size_code = fallback_code.split("-", 1)[0]
    return size_code or fallback_code


def derive_title(title_value: Any, description_value: Any, code: str) -> str:
    title = clean(title_value)
    if title:
        return title
    description = clean(description_value)
    if description:
        return description.split("-")[0].strip()
    return code


def normalize_for_match(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value.lower())
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def resolve_site_category_from_excel(category_value: str) -> str:
    normalized = normalize_for_match(clean(category_value))
    if "ειδη σπιτιου" in normalized:
        return "Είδη Σπιτιού"
    if "γλαστρες" in normalized:
        return "Γλάστρες"
    return "Επαγγελματικός Εξοπλισμός"


def is_green_description_cell(cell: Any) -> bool:
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type != "solid":
        return False

    fg_color = getattr(fill, "fgColor", None)
    if not fg_color or getattr(fg_color, "type", None) != "rgb":
        return False

    rgb = (fg_color.rgb or "").upper()
    return rgb.endswith(GREEN_FILL_HEX_SUFFIX)


def build_rows(worksheet: Any) -> list[dict[str, Any]]:
    current_group = ""
    current_family = ""
    seen_codes: set[str] = set()
    rows: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=3, max_col=COL_LIFESTYLE_END + 1, values_only=False):
        marker_group = clean(row[COL_GROUP_MARKER].value)
        marker_family = clean(row[COL_FAMILY_MARKER].value)
        code = clean(row[COL_CODE].value)

        # Keep the latest marker context, whether marker values appear on
        # dedicated marker rows or directly on data rows.
        if marker_group:
            current_group = marker_group
        if marker_family:
            current_family = marker_family

        # Marker-only rows carry context but no variant data.
        if not code:
            continue
        if not is_green_description_cell(row[COL_DESCRIPTION]):
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)

        title = derive_title(row[COL_TITLE_GR].value, row[COL_DESCRIPTION].value, code)
        en_slug = clean(row[COL_TITLE_EN_SLUG].value)
        size_code = code.split("-", 1)[0] if "-" in code else code
        packshot = clean(row[COL_PACKSHOT].value)
        additional_images = [
            clean(row[index].value)
            for index in range(COL_LIFESTYLE_START, COL_LIFESTYLE_END + 1)
            if clean(row[index].value)
        ]

        rows.append(
            {
                "group_root": current_group,
                "family_indicator": current_family,
                "group_code": extract_group_code(current_group, code),
                "title": title,
                "title_en_slug": en_slug,
                "category": resolve_site_category_from_excel(clean(row[COL_CATEGORY].value)),
                "size_code": size_code,
                "variant": {
                    "code": code,
                    "description": clean(row[COL_DESCRIPTION].value),
                    "color": clean(row[COL_COLOR].value),
                    "image_url": packshot,
                    "pack": clean(row[COL_PACK].value),
                    "excel_ar": clean(row[COL_EXCEL_AR].value),
                },
                "additional_images": additional_images,
            }
        )

    return rows


def build_grouped_products(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        # Group strictly by the Excel W marker (group_root).
        key = row["group_root"] or row["group_code"]
        grouped[key].append(row)

    products: list[dict[str, Any]] = []

    for group_key, product_rows in grouped.items():
        group_code = product_rows[0]["group_code"]
        title = product_rows[0]["title"]
        product_id = group_code

        variants_by_size: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in product_rows:
            variants_by_size[row["size_code"]].append(row["variant"])

        sizes = []
        for size_code, variants in sorted(variants_by_size.items(), key=lambda item: item[0]):
            colors_count = len({clean(variant["color"]) for variant in variants if clean(variant["color"])})
            specs_source = next(
                (
                    variant
                    for variant in variants
                    if parse_specs_from_text(clean(variant.get("description"))).get("has_specs")
                ),
                variants[0] if variants else {},
            )
            sizes.append(
                {
                    "size_label": size_code,
                    "size_code": size_code,
                    "variants": variants,
                    "colors_count": colors_count,
                    "specs": parse_specs_from_text(clean(specs_source.get("description"))),
                }
            )

        representative_image = ""
        for size in sizes:
            for variant in size["variants"]:
                image_url = clean(variant.get("image_url"))
                if image_url:
                    representative_image = image_url
                    break
            if representative_image:
                break

        variants_count = sum(len(size["variants"]) for size in sizes)
        category_counts = Counter(row["category"] for row in product_rows if clean(row.get("category")))
        resolved_category = (
            category_counts.most_common(1)[0][0]
            if category_counts
            else "Επαγγελματικός Εξοπλισμός"
        )
        products.append(
            {
                "id": product_id,
                "family_indicator": clean(product_rows[0]["family_indicator"]),
                "group_root": clean(product_rows[0]["group_root"]),
                "title": title,
                "category": resolved_category,
                "representative_image": representative_image,
                "sizes": sizes,
                "sizes_count": len(sizes),
                "variants_count": variants_count,
            }
        )

    products.sort(key=lambda product: (product["group_root"], product["title"], product["id"]))
    return products


def parse_specs_from_text(text: str) -> dict[str, Any]:
    normalized = clean(text).lower()
    if not normalized:
        return {
            "liters": None,
            "width": None,
            "depth": None,
            "box_height": None,
            "diameter": None,
            "height": None,
            "has_specs": False,
        }

    liters_match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:lt|l)\b", normalized, flags=re.IGNORECASE)
    box_3d_match = re.search(
        r"(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)(?:\s*h)?",
        normalized,
        flags=re.IGNORECASE,
    )
    # 2D sizes like stickers: 8,5x33 cm (no depth)
    box_2d_match = None
    if not box_3d_match:
        box_2d_match = re.search(
            r"(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)(?:\s*(?:cm|χιλ|mm)\b|\b)",
            normalized,
            flags=re.IGNORECASE,
        )

    dimensions_match = re.search(
        r"(?:^|[^a-z0-9])d\s*([0-9]+(?:[.,][0-9]+)?)\s*x\s*([0-9]+(?:[.,][0-9]+)?)\s*h?",
        normalized,
        flags=re.IGNORECASE,
    ) or re.search(
        r"ø\s*([0-9]+(?:[.,][0-9]+)?)\s*x\s*h?\s*([0-9]+(?:[.,][0-9]+)?)",
        normalized,
        flags=re.IGNORECASE,
    )

    liters = liters_match.group(1).replace(",", ".") if liters_match else None
    explicit_diameter = dimensions_match.group(1).replace(",", ".") if dimensions_match else None
    explicit_height = dimensions_match.group(2).replace(",", ".") if dimensions_match else None

    box_width = box_3d_match.group(1).replace(",", ".") if box_3d_match else None
    box_depth = box_3d_match.group(2).replace(",", ".") if box_3d_match else None
    box_height = box_3d_match.group(3).replace(",", ".") if box_3d_match else None

    # If diameter syntax exists (d..x..h or ø..x..), do not treat it as width-height.
    if explicit_diameter:
        box_2d_match = None

    two_d_width = box_2d_match.group(1).replace(",", ".") if box_2d_match else None
    two_d_height = box_2d_match.group(2).replace(",", ".") if box_2d_match else None

    # Some records encode round products as 8x16x16h; normalize to Ø8 x H16.
    looks_like_round_duplicate = (
        not explicit_diameter
        and box_width
        and box_depth
        and box_height
        and box_depth == box_height
    )

    width = None
    depth = None
    height = explicit_height
    parsed_box_height = None
    diameter = explicit_diameter

    if looks_like_round_duplicate:
        diameter = explicit_diameter or box_width
        height = explicit_height or box_depth
    elif box_3d_match:
        width = box_width
        depth = box_depth
        parsed_box_height = box_height
    elif box_2d_match:
        width = two_d_width
        # For flat/2D specs, treat second value as height so UI can display W x H.
        height = two_d_height

    has_specs = any([liters, width, depth, parsed_box_height, diameter, height])
    return {
        "liters": liters,
        "width": width,
        "depth": depth,
        "box_height": parsed_box_height,
        "diameter": diameter,
        "height": height,
        "has_specs": has_specs,
    }


def build_additional_images(rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    for row in rows:
        code = row["variant"]["code"]
        images = row["additional_images"]
        if not images:
            continue
        deduped = list(dict.fromkeys(images))
        mapping[code] = deduped

    return dict(sorted(mapping.items(), key=lambda item: item[0]))


def is_url_reachable(url: str, timeout_seconds: float) -> bool:
    if not url or not url.lower().startswith(("http://", "https://")):
        return False

    headers = {"User-Agent": "viomes-catalog-json-generator/1.0"}
    for method in ("HEAD", "GET"):
        request = Request(url, headers=headers, method=method)
        try:
            with urlopen(request, timeout=timeout_seconds) as response:
                status = getattr(response, "status", None) or response.getcode()
                return 200 <= int(status) < 400
        except HTTPError as exc:
            # Some servers reject HEAD even if GET would work.
            if method == "HEAD" and exc.code in (405, 501):
                continue
            return False
        except (URLError, TimeoutError):
            return False
        except Exception:
            return False

    return False


def filter_unreachable_additional_images(
    mapping: dict[str, list[str]],
    timeout_seconds: float,
    workers: int,
) -> tuple[dict[str, list[str]], int, int]:
    all_urls = sorted({url for urls in mapping.values() for url in urls})
    if not all_urls:
        return mapping, 0, 0

    reachable_by_url: dict[str, bool] = {}
    worker_count = max(1, min(workers, len(all_urls)))
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(is_url_reachable, url, timeout_seconds): url
            for url in all_urls
        }
        for future in as_completed(futures):
            url = futures[future]
            reachable_by_url[url] = future.result()

    filtered: dict[str, list[str]] = {}
    total_urls = 0
    removed_urls = 0
    for code, urls in mapping.items():
        total_urls += len(urls)
        valid_urls = [url for url in urls if reachable_by_url.get(url, False)]
        removed_urls += len(urls) - len(valid_urls)
        if valid_urls:
            filtered[code] = valid_urls

    return dict(sorted(filtered.items(), key=lambda item: item[0])), total_urls, removed_urls


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate catalog JSON files from XLSX.")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Path to the source workbook. Defaults to the first .xlsx in this folder.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--products-out",
        type=Path,
        default=Path("src/data/products-grouped.json"),
        help="Output path for products-grouped JSON.",
    )
    parser.add_argument(
        "--additional-out",
        type=Path,
        default=Path("src/data/additional-images.json"),
        help="Output path for additional-images JSON.",
    )
    parser.add_argument(
        "--skip-url-validation",
        action="store_true",
        help="Skip URL reachability checks for additional images.",
    )
    parser.add_argument(
        "--url-timeout",
        type=float,
        default=6.0,
        help="Timeout in seconds for each additional-image URL check.",
    )
    parser.add_argument(
        "--url-workers",
        type=int,
        default=24,
        help="Parallel workers for additional-image URL checks.",
    )
    return parser.parse_args()


def resolve_xlsx_path(explicit_path: Path | None) -> Path:
    if explicit_path:
        return explicit_path

    candidates = sorted(
        path for path in Path("src/data").glob("*.xlsx") if not path.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError("No .xlsx file found in src/data")
    return candidates[0]


def main() -> None:
    args = parse_args()
    xlsx_path = resolve_xlsx_path(args.xlsx)

    workbook = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]

    rows = build_rows(worksheet)
    products = build_grouped_products(rows)
    additional_images = build_additional_images(rows)
    total_additional_urls = sum(len(urls) for urls in additional_images.values())
    removed_additional_urls = 0
    if not args.skip_url_validation:
        additional_images, total_additional_urls, removed_additional_urls = filter_unreachable_additional_images(
            additional_images,
            timeout_seconds=args.url_timeout,
            workers=args.url_workers,
        )

    products_payload = {
        "source_file": xlsx_path.name,
        "products_count": len(products),
        "products": products,
    }

    args.products_out.parent.mkdir(parents=True, exist_ok=True)
    args.additional_out.parent.mkdir(parents=True, exist_ok=True)

    args.products_out.write_text(
        json.dumps(products_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    args.additional_out.write_text(
        json.dumps(additional_images, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Wrote {args.products_out} ({len(products)} products)")
    print(f"Wrote {args.additional_out} ({len(additional_images)} variant image groups)")
    if args.skip_url_validation:
        print("Skipped additional-image URL validation")
    else:
        print(
            f"Validated additional-image URLs: removed {removed_additional_urls} "
            f"of {total_additional_urls} links"
        )


if __name__ == "__main__":
    main()
